
###############################DEPENDENCIES##############################################
import sys
import requests
from bs4 import BeautifulSoup
import os, os.path
import datetime
from openpyxl import Workbook
import xlsxwriter
########################################################################################



################## This function checks and creates image folder #######################
def image_folder():
    folder_name = "Images"
    if not os.path.exists(folder_name):
        os.makedirs(folder_name)
    __location__ = os.path.realpath(os.path.join(os.getcwd(), os.path.dirname(__file__)))
    location = os.path.join(__location__, folder_name)
    return location, __location__
#########################################################################################



############ This funtion fetches images from webpage and stores in Image folder ########
def fetching_images(path):
 print("Image Fetching Started ...")
 data_image=[]
 count_found=0
 count_not_founded=0
 try:
    location=image_folder()
    req = requests.get(path)
    data = req.text
    soup = BeautifulSoup(data, "lxml")
    for link in soup.find_all('img'):
        image = link.get("src")
        if image[0:4]=="http":
            image=image
        else:
            image = "http:" + image
        if image.find("?"):
            image = image[:image.find("?")]
        image_name = os.path.split(image)[1]
        data_image.append(image)
        req_image = requests.get(image)
        count_found=count_found+1
        with open(os.path.join(location[0],image_name), "wb") as f:
           f.write(req_image.content)
        with open(os.path.join(location[1],"List_of_all_URLs.txt"),"a") as p:
            p.write(image+"\n")
 except ValueError:
     print("Incorrect URL for specific image")
     count_not_founded=count_not_founded+1
 count_downloadable=count_found-count_not_founded

 if count_downloadable < 0:
    count_downloadable=0
 else:
    count_downloadable=count_downloadable
 write_data(data_image,count_not_founded)
 print("Total Images Found: ",count_found)
 #print("Total downloadeable images: ",count_downloadable)
 #print("Error in URL's: ",count_not_founded)
########################################################################################



############## This function creates new excel file ####################################
def create_book(path_link):
    workbook = xlsxwriter.Workbook(path_link)
    worksheet = workbook.add_worksheet('data')
    worksheet.write(0, 0, 'List of all URLs')
    workbook.close()
########################################################################################


############### This function write URL's to excel file ################################
def write_data(data,count):
 try:
    start_data=2
    path_link="last_search_result.xlsx"
    if (not os.path.isfile(path_link)):
        create_book(path_link)
    wb = Workbook()
    sheet1 = wb.active
    sheet1.title = 'data'
    for i in range(1, len(data)+1-count):
        _ = sheet1.cell(column=1, row=i, value=data[i-1])
    wb.save(filename=path_link)
    wb.close()

 except Exception as e:
    print("Close Excel File !!!! \nFirst close excel file then \nClick Ok or Cancel to continue")
##########################################################################################



############### This function checks the version of python ###############################
def version_control():
    py3=False
    if sys.version_info[0] > 2:
        py3=True
    return py3
########################################################################################



############### This function checks the continuation of program ###############################
def continue_exit(py3):
    if py3:
        decision = input("Do you want to use this program again !!! Press Y to continue and N to exit ").lower()
    else:
        decision = raw_input("Do you want to use this program again !!! Press Y to continue and N to exit ").lower()
    if decision == "y":
        main()
    elif decision == "n":
        sys.exit()
    else:
        print("Incorrect input try again !!!")
########################################################################################


################# This is main function, which calls other functions ####################
def main():
    py3=version_control()
    if py3:
        path = input("Enter the path you want to fetch images from: ")
    else:
        path = raw_input("Enter the path you want to fetch images from: ")
    if path=="":
        print("Nothing entered!!!!")
        continue_exit(py3)
    start = datetime.datetime.now()
    fetching_images(path)
    finish = datetime.datetime.now()
    print("Execution time (hh:mm:ss:ms): ", finish-start)
    while 1:
            continue_exit(py3)
########################################################################################



# Program execution starts here
if __name__=="__main__":
    app=main()